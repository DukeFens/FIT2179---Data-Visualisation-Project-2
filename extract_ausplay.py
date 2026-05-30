import requests
import openpyxl
import io

url = "https://ascwrstorageprod001.ausport.gov.au/assets/1123268-PnnmSxp8rFCjHBk2GdwAOw.xlsx?sv=2025-11-05&st=2026-05-28T12%3A48%3A26Z&se=2026-05-28T20%3A48%3A26Z&sr=b&sp=r&sig=qF3LnelvN9tSUrYLbmnMQwjQogDUJj8Ek8wOjTygqfY%3D"
print("Downloading...")
resp = requests.get(url)
print("Parsing Excel with openpyxl...")
wb = openpyxl.load_workbook(filename=io.BytesIO(resp.content), data_only=True)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    if 'data' in sheet_name.lower() or 'trend' in sheet_name.lower() or 'sport' in sheet_name.lower():
        sheet = wb[sheet_name]
        print(f"\nSheet {sheet_name} head:")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i > 10: break
            print(row)
        break
